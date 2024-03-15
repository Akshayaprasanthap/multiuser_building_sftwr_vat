
from django.http import HttpResponseNotFound, JsonResponse
from django.shortcuts import render, redirect
from .models import *
from django.contrib import messages
from django.contrib.auth.models import User, auth
from django.utils.crypto import get_random_string
import random
from django.conf import settings
from django.core.mail import send_mail
from django.utils import timezone
from .models import Company
from django.shortcuts import render, redirect
from django.http import JsonResponse


from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from io import BytesIO
from xhtml2pdf import pisa
from django.http.response import JsonResponse, HttpResponse


def home(request):
  return render(request, 'home.html')

def login(request):
  return render(request, 'login.html')

def forgot_password(request):
  return render(request, 'forgot_password.html')

def cmp_register(request):
  return render(request, 'cmp_register.html')

def cmp_details(request,id):
  context = {'id':id}
  return render(request, 'cmp_details.html', context)

def emp_register(request):
  return render(request, 'emp_register.html')

def register_company(request):
  if request.method == 'POST':
    fname = request.POST['fname']
    lname = request.POST['lname']
    email = request.POST['email']
    uname = request.POST['uname']
    phno = request.POST['phno']
    passw = request.POST['pass']
    cpass = request.POST['cpass']
    rfile = request.FILES.get('rfile')

    if passw == cpass:
      if CustomUser.objects.filter(username = uname).exists():
        messages.info(request, 'Sorry, Username already exists !!')
        return redirect('cmp_register')

      elif not CustomUser.objects.filter(email = email).exists():
        user_data = CustomUser.objects.create_user(first_name = fname, last_name = lname, username = uname, email = email, password = passw, is_company = 1)
        cmp = Company( contact = phno, user = user_data, profile_pic = rfile)
        cmp.save()
        return redirect('cmp_details',user_data.id)

      else:
        messages.info(request, 'Sorry, Email already exists !!')
        return redirect('cmp_register')
      
    messages.info(request, 'Sorry, Passwords must match !!')
    return render(request,'cmp_register.html')
  
def register_company_details(request,id):
  if request.method == 'POST':
    cname = request.POST['cname']
    address = request.POST['address']
    city = request.POST['city']
    state = request.POST['state']
    country = request.POST['country']
    pincode = request.POST['pincode']
    pannumber = request.POST['pannumber']
    gsttype = request.POST['gsttype']
    gstno = request.POST['gstno']

    code=get_random_string(length=6)

    usr = CustomUser.objects.get(id = id)
    cust = Company.objects.get(user = usr)
    cust.company_name = cname
    cust.address = address
    cust.city = city
    cust.state = state
    cust.company_code = code
    cust.country = country
    cust.pincode = pincode
    cust.pan_number = pannumber
    cust.gst_type = gsttype
    cust.gst_no = gstno
    cust.save()
    return redirect('login')

def register_employee(request):
  if request.method == 'POST':
    fname = request.POST['fname']
    lname = request.POST['lname']
    email = request.POST['email']
    uname = request.POST['uname']
    phno = request.POST['phno']
    passw = request.POST['pass']
    cpass = request.POST['cpass']
    ccode = request.POST['ccode']
    rfile = request.FILES.get('rfile')

    if not Company.objects.filter(company_code = ccode).exists():
      messages.info(request, 'Sorry, Company Code is Invalid !!')
      return redirect('emp_register')
    
    cmp = Company.objects.get(company_code = ccode)
    emp_names = Employee.objects.filter(company = cmp).values_list('user',flat=True)
    for e in emp_names:
       usr = CustomUser.objects.get(id=e)
       if str(fname).lower() == (usr.first_name ).lower() and str(lname).lower() == (usr.last_name).lower():
        messages.info(request, 'Sorry, Employee With this name already exits, try adding an initial !!')
        return redirect('emp_register')
    
    if passw == cpass:
      if CustomUser.objects.filter(username = uname).exists():
        messages.info(request, 'Sorry, Username already exists !!')
        return redirect('emp_register')

      elif not CustomUser.objects.filter(email = email).exists():
        user_data = CustomUser.objects.create_user(first_name = fname, last_name = lname, username = uname, email = email, password = passw)
        emp = Employee(user = user_data, company = cmp, profile_pic = rfile, contact=phno)
        emp.save()
        return redirect('login')

      else:
        messages.info(request, 'Sorry, Email already exists !!')
        return redirect('emp_register')
      
    messages.info(request, 'Sorry, Passwords must match !!')
    return render(request,'emp_register.html')
  
def change_password(request):
  if request.method == 'POST':
    email= request.POST.get('email')
    if not CustomUser.objects.filter(email=email).exists():
      messages.success(request,'No user found with this email !!')
      return redirect('forgot_password')
    
    else:
      otp = random.randint(100000, 999999)
      usr = CustomUser.objects.get(email=email)
      usr.set_password(str(otp))
      usr.save()

      subject = 'Password Reset Mail'
      message = f'Hi {usr.first_name} {usr.last_name}, Your Otp for password reset is {otp}'
      email_from = settings.EMAIL_HOST_USER
      recipient_list = [email ]
      send_mail(subject, message, email_from, recipient_list)
      messages.info(request,'Password reset mail sent !!')
      return redirect('forgot_password')

def user_login(request):
  if request.method == 'POST':
    email = request.POST['email']
    cpass = request.POST['pass']

    try:
      usr = CustomUser.objects.get(email=email)
      log_user = auth.authenticate(username = usr.username, password = cpass)
      if log_user is not None:
        if usr.is_company == 1:
          auth.login(request, log_user)
          return redirect('dashboard')
        else:
          emp = Employee.objects.get(user=usr)
          if emp.is_approved == 0:
            messages.info(request,'Employee is not Approved !!')
            return redirect('login')
          else:
            auth.login(request, log_user)
            return redirect('dashboard')
      messages.info(request,'Invalid Login Details !!')
      return redirect('login')
    
    except:
        messages.info(request,'Employee do not exist !!')
        return redirect('login')
    

def dashboard(request):
  context = {'usr':request.user}
  return render(request, 'dashboard.html', context)

def logout(request):
  auth.logout(request)
  return redirect('/')

def cmp_profile(request):
  cmp = Company.objects.get(user = request.user)
  context = {'usr':request.user, 'cmp':cmp}
  return render(request,'cmp_profile.html',context)

def load_edit_cmp_profile(request):
  cmp = Company.objects.get(user = request.user)
  context = {'usr':request.user, 'cmp':cmp}
  return render(request,'cmp_profile_edit.html',context)

def edit_cmp_profile(request):
  cmp =  Company.objects.get(user = request.user)
  if request.method == 'POST':
    email = request.POST['email']
    current_email = cmp.user.email
    if email != current_email:
      if CustomUser.objects.filter(email=email).exists():
        messages.info(request,'Email Already in Use')
        return redirect('load_edit_cmp_profile')

    cmp.company_name = request.POST['cname']
    cmp.user.email = request.POST['email']
    cmp.user.first_name = request.POST['fname']
    cmp.user.last_name = request.POST['lname']
    cmp.contact = request.POST['phno']
    cmp.address = request.POST['address']
    cmp.city = request.POST['city']
    cmp.state = request.POST['state']
    cmp.country = request.POST['country']
    cmp.pincode = request.POST['pincode']
    cmp.pan_number = request.POST['pan']
    cmp.gst_type = request.POST['gsttype']
    cmp.gst_no = request.POST['gstnoval']
    old=cmp.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      cmp.profile_pic=old
    else:
      cmp.profile_pic=new
    
    cmp.save() 
    cmp.user.save() 
    return redirect('cmp_profile') 
  
def emp_profile(request):
  emp = Employee.objects.get(user=request.user)
  context = {'usr':request.user, 'emp':emp}
  return render(request,'emp_profile.html',context)

def load_edit_emp_profile(request):
  emp = Employee.objects.get(user=request.user)
  context = {'usr':request.user, 'emp':emp}
  return render(request,'emp_profile_edit.html',context)

def edit_emp_profile(request):
  emp =  Employee.objects.get(user = request.user)
  if request.method == 'POST':
    email = request.POST['email']
    current_email = emp.user.email
    if email != current_email:
      if CustomUser.objects.filter(email=email).exists():
        messages.info(request,'Email Already in Use')
        return redirect('load_edit_emp_profile')

    emp.user.email = request.POST['email']
    emp.user.first_name = request.POST['fname']
    emp.user.last_name = request.POST['lname']
    emp.contact = request.POST['phno']
    old=emp.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      emp.profile_pic=old
    else:
      emp.profile_pic=new
    
    emp.save() 
    emp.user.save() 
    return redirect('emp_profile') 

def load_staff_request(request):
  cmp = Company.objects.get(user = request.user)
  emp = Employee.objects.filter(company = cmp, is_approved = 0)
  context = {'usr':request.user, 'emp':emp, 'cmp':cmp}
  return render(request,'staff_request.html',context)

def load_staff_list(request):
  cmp = Company.objects.get(user = request.user)
  emp = Employee.objects.filter(company = cmp, is_approved = 1)
  context = {'usr':request.user, 'emp':emp, 'cmp':cmp}
  return render(request,'staff_list.html',context)

def accept_staff(request,id):
  emp = Employee.objects.get(id=id)
  emp.is_approved = 1
  emp.save()
  messages.info(request,'Employee Approved !!')
  return redirect('load_staff_request')

def reject_staff(request,id):
  Employee.objects.get(id=id).delete()
  messages.info(request,'Employee Deleted !!')
  return redirect('load_staff_request')

def item_list(request):
  if request.user.is_company:
    itm = Item.objects.filter(company = request.user.company)
  else:
    itm = Item.objects.filter(company = request.user.employee.company)
  
  if itm:
    fitm = itm[0]
    ftrans = Transactions.objects.filter(item = fitm)
    context = {'itm':itm, 'usr':request.user, 'fitm':fitm, 'ftrans':ftrans}
  else:
        context = {'itm':itm, 'usr':request.user}
  return render(request,'item_list.html',context)

def load_item_create(request):
  tod = timezone.now().date().strftime("%Y-%m-%d")
  return render(request,'item_create.html',{'tod':tod, 'usr':request.user})

def item_create(request):
  if request.method=='POST':
    itm_type = request.POST.get('itm_type')
    if itm_type:
      type = 'Service'
    else:
      type = 'Goods'
    itm_name = request.POST.get('name')
    itm_hsn = request.POST.get('hsn')
    itm_unit = request.POST.get('unit')
    itm_vat = request.POST.get('vat')
    taxable_result = request.POST.get('taxable_result')
    itm_sale_price = request.POST.get('sale_price')
    itm_purchase_price = request.POST.get('purchase_price')
    stock_in_hand = request.POST.get('stock_in_hand')
    if stock_in_hand == '' or None :
      stock_in_hand = 0
    itm_at_price = request.POST.get('at_price')
    if itm_at_price == '' or None:
      itm_at_price = 0
    itm_date = request.POST.get('itm_date')
    
    item = Item(user = request.user,
                itm_type = type,
                itm_name = itm_name,
                itm_hsn = itm_hsn,
                itm_unit = itm_unit,
                itm_vat = itm_vat,
                itm_taxable = taxable_result,
                itm_sale_price = itm_sale_price,
                itm_purchase_price = itm_purchase_price,
                itm_stock_in_hand = stock_in_hand,
                itm_at_price = itm_at_price,
                itm_date = itm_date)
    item.save()

    trans = Transactions(user = request.user, item = item, trans_type = 'Stock Open', trans_date = itm_date, trans_qty = 0, trans_current_qty = stock_in_hand, 
                         trans_adjusted_qty = stock_in_hand, trans_price = itm_at_price)

    if request.user.is_company:
      item.company = request.user.company
      trans.company = request.user.company

    else:
      item.company = request.user.employee.company
      trans.company = request.user.employee.company
 
    item.save()
    trans.save()

    if request.POST.get('save_and_next'):
      return redirect('load_item_create')
    elif request.POST.get('save'):
      return redirect('item_list')
    





def party_list(request):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)
  
  if party:
    fparty = party[0]
    ftrans = Transactions_party.objects.filter(party = fparty)
    context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans}
  else:
        context = {'party':party, 'usr':request.user}
  return render(request,'parties_list.html',context)


def load_party_create(request):
  tod = timezone.now().date().strftime("%Y-%m-%d")
  return render(request,'add_parties.html',{'tod':tod, 'usr':request.user})

from django.db import IntegrityError
from django.shortcuts import render, redirect

def addNewParty(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company

    error_message = None  # Initialize error message

    if request.method == 'POST':
        party_name = request.POST['partyname']
        trn_no = request.POST['trn_no']
        contact = request.POST['contact']
        trn_type = request.POST['trn_type']
        state = request.POST.get('state')
        address = request.POST['address']
        email = request.POST['email']
        opening_stock=request.POST['opening_stock']
        at_price=request.POST['at_price']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']

        try:
            # Check if a party with the same trn_no already exists
            if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
                error_message = 'An error occurred while processing your request. TRN number already exists. Please enter a unique TRN number.'
            # Check if a party with the same email already exists
            elif Party.objects.filter(email=email, company=cmp).exists():
                error_message = 'An error occurred while processing your request. Email already exists. Please enter a unique email address.'
            else:
                part = Party(party_name=party_name, trn_no=trn_no, contact=contact, trn_type=trn_type, state=state,
                             address=address, email=email, openingbalance=openingbalance,opening_stock=opening_stock,at_price=at_price ,payment=payment,
                             current_date=current_date, End_date=End_date, additionalfield1=additionalfield1,
                             additionalfield2=additionalfield2, additionalfield3=additionalfield3)

                if request.user.is_company:
                    part.company = request.user.company
                else:
                    part.company = request.user.employee.company

                part.save()

                trans = Transactions_party(user=request.user, trans_type='Opening Balance', trans_number=trn_no,
                                           trans_date=current_date, total=openingbalance, balance=openingbalance, party=part)

                if request.user.is_company:
                    trans.company = request.user.company
                else:
                    trans.company = request.user.employee.company

                trans.save()

                tr_history = PartyTransactionHistory(party=part, Transactions_party=trans, action="CREATED")
                tr_history.save()

                if request.POST.get('save_and_next'):
                    return redirect('load_party_create')
                elif request.POST.get('save'):
                    return redirect('party_list')

        except IntegrityError:
            # Specific error message for duplicate TRN number
            error_message = 'An error occurred while processing your request. Please try again.'

    return render(request, 'add_parties.html', {'error_message': error_message})









# def addNewParty(request):
#     if request.user.is_company:
#         cmp = request.user.company
#     else:
#         cmp = request.user.employee.company
#         print(cmp)

#     if request.method == 'POST':
#         party_name = request.POST['partyname']
#         trn_no = request.POST['trn_no']
#         contact = request.POST['contact']
#         trn_type = request.POST['trn_type']
#         state = request.POST.get('state')
#         address = request.POST['address']
#         email = request.POST['email']
#         openingbalance = request.POST.get('balance', '')
#         payment = request.POST.get('paymentType', '')
#         current_date = request.POST['currentdate']
#         End_date = request.POST.get('enddate', None)
#         additionalfield1 = request.POST['additionalfield1']
#         additionalfield2 = request.POST['additionalfield2']
#         additionalfield3 = request.POST['additionalfield3']

#         part = Party(party_name=party_name, trn_no=trn_no, contact=contact, trn_type=trn_type, state=state,
#                      address=address, email=email, openingbalance=openingbalance, payment=payment,
#                      current_date=current_date, End_date=End_date, additionalfield1=additionalfield1,
#                      additionalfield2=additionalfield2, additionalfield3=additionalfield3)

#         if request.user.is_company:
#             part.company = request.user.company
#         else:
#             part.company = request.user.employee.company

#         part.save()

#         trans = Transactions_party(user=request.user, trans_type='Opening Balance', trans_number=trn_no,
#                                    trans_date=current_date, total=openingbalance, balance=openingbalance, party=part)

#         if request.user.is_company:
#             trans.company = request.user.company
#         else:
#             trans.company = request.user.employee.company

#         trans.save()

#         tr_history = PartyTransactionHistory(party=part,Transactions_party=trans, action="CREATED")
#         tr_history.save()

#         if request.POST.get('save_and_next'):
#             return redirect('load_party_create')
#         elif request.POST.get('save'):
#             return redirect('party_list')

#     return render(request, 'add_parties.html')









# def addNewParty(request):
#     if request.user.is_company:
#         cmp = request.user.company
#     else:
#         cmp = request.user.employee.company
#         print(cmp)
#     if request.method == 'POST':
#       party_name = request.POST['partyname']
#       trn_no = request.POST['trn_no']
#       contact = request.POST['contact']
#       trn_type = request.POST['trn_type']
#       state = request.POST.get('state')
#       address = request.POST['address']
#       email = request.POST['email']
#       openingbalance = request.POST.get('balance', '')
#       payment = request.POST.get('paymentType', '')
#       current_date = request.POST['currentdate']
#       End_date = request.POST.get('enddate', None)
#       additionalfield1 = request.POST['additionalfield1']
#       additionalfield2 = request.POST['additionalfield2']
#       additionalfield3 = request.POST['additionalfield3']

#       part = Party(party_name=party_name, trn_no=trn_no,contact=contact,trn_type=trn_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
#                       current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3)
#       part.save()

#       # return JsonResponse({'status':True})
#       trans = Transactions_party(user = request.user,trans_type ='Opening Balance',trans_number=trn_no,trans_date =current_date,total=openingbalance, balance=openingbalance,party=part)
#       tr_history = PartyTransactionHistory(party=part,
#                                               Transactions_party=trans,      
#                                               action="CREATED",)
#       tr_history.save()
#     if request.user.is_company:
#       part.company = request.user.company
#       trans.company = request.user.company

#     else:
#       part.company = request.user.employee.company
#       trans.company = request.user.employee.company
 
#     part.save()
#     trans.save()

#     if request.POST.get('save_and_next'):
#       return redirect('load_party_create')
#     elif request.POST.get('save'):
#       return redirect('party_list')
    



def view_party(request,id):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)

  fparty = Party.objects.get(id=id)
  ftrans = Transactions_party.objects.filter(party = fparty)
  context = {'party':party, 'usr':request.user, 'fparty':fparty, 'ftrans':ftrans}
  return render(request,'parties_list.html',context)



def edit_party(request,id):
  if request.user.is_company:
    party = Party.objects.filter(company = request.user.company)
  else:
    party = Party.objects.filter(company = request.user.employee.company)

  getparty=Party.objects.get(id=id)
  parties=Party.objects.filter(user=request.user)
  ftrans = Transactions_party.objects.filter(party = getparty)
  return render(request, 'edit_party.html',{'usr':request.user,'party':party,'getparty':getparty,'parties':parties,'ftrans':ftrans})

# from django.shortcuts import get_object_or_404

# def edit_party(request, id):
#   if request.user.is_company:
#       party = Party.objects.filter(company=request.user.company)
#   else:
#       party = Party.objects.filter(company=request.user.employee.company)

#   try:
#       getparty = get_object_or_404(Party, id=id)
#       ftrans = Transactions_party.objects.filter(party=getparty)
#       parties = Party.objects.filter(user=request.user)
#       return render(request, 'edit_party.html', {'usr': request.user, 'party': party, 'getparty': getparty, 'parties': parties, 'ftrans': ftrans})
#   except Party.DoesNotExist:
#         # Handle the case when the party with the specified ID does not exist
#       return HttpResponseNotFound("Party does not exist")




# from django.shortcuts import render, redirect
# from .models import Party, Transactions_party


from datetime import datetime
from django.shortcuts import get_object_or_404

def edit_saveparty(request, id):
    company = request.user.company if request.user.is_company else request.user.employee.company
    party_qs = Party.objects.filter(company=company)
    getparty = get_object_or_404(Party, id=id, company=company)
    
    trans = None  # Initialize trans here

    if request.method == 'POST':
        # Update party details
        getparty.party_name = request.POST.get('partyname')
        getparty.trn_no = request.POST.get('trn_no')
        getparty.contact = request.POST['contact']
        getparty.trn_type = request.POST['trn_type']
        getparty.state = request.POST['state']
        getparty.address = request.POST['address']
        getparty.email = request.POST['email']
        getparty.openingbalance = request.POST['balance']
        getparty.payment = request.POST.get('paymentType')
        # getparty.current_date = request.POST['currentdate']
        getparty.current_date = datetime.now().strftime('%Y-%m-%d')
        getparty.additionalfield1 = request.POST['additionalfield1']
        getparty.additionalfield2 = request.POST['additionalfield2']
        getparty.additionalfield3 = request.POST['additionalfield3']

        # Save the party changes
        getparty.save()

        # Check if a transaction with the same party_id exists
        existing_transaction = Transactions_party.objects.filter(party_id=id).first()

        # Update or create a new transaction
        if existing_transaction:
            existing_transaction.trans_type = 'Opening Balance'
            existing_transaction.trans_number = getparty.trn_no
            existing_transaction.trans_date = getparty.current_date
            existing_transaction.total = getparty.openingbalance
            existing_transaction.balance = getparty.openingbalance
            existing_transaction.save()

            # Check and update the transaction history action to "UPDATED"
            party_history_entry = PartyTransactionHistory.objects.filter(
                party=getparty, Transactions_party=existing_transaction
            ).first()
            if party_history_entry:
                party_history_entry.action = "UPDATED"
                party_history_entry.save()

        else:
            # Create and save a new transaction
            trans = Transactions_party(
                user=request.user,
                trans_type='Opening Balance',
                trans_number=getparty.trn_no,
                trans_date=getparty.current_date,
                total=getparty.openingbalance,
                balance=getparty.openingbalance,
                party=getparty,
                company=company
            )
            trans.save()

            # Save the transaction history entry with "CREATED" action
            tr_history = PartyTransactionHistory(
                party=getparty,
                Transactions_party=trans,
                action="UPDATED"
            )
            tr_history.save()

        return redirect('view_party', id=getparty.id)

    return render(request, 'edit_party.html', {'getparty': getparty, 'party': party_qs, 'usr': request.user})




# tr_history = PartyTransactionHistory(party=getparty,Transactions_party=trans,action="UPDATED")
#             print(tr_history,'000000000000000000')
#             tr_history.save()


# Remove has_changed function if not used

# from django.utils import timezone

# def edit_saveparty(request, id):
#     if request.user.is_company:
#         party_qs = Party.objects.filter(company=request.user.company)
#     else:
#         party_qs = Party.objects.filter(company=request.user.employee.company)

#     getparty = Party.objects.get(id=id)

#     if request.method == 'POST':
#         # Capture old data before making changes
#         old_data = {
#             'party_name': getparty.party_name,
#             'trn_no': getparty.trn_no,
#             'contact': getparty.contact,
#             'trn_type': getparty.trn_type,
#             'state': getparty.state,
#             'address': getparty.address,
#             'email': getparty.email,
#             'openingbalance': getparty.openingbalance,
#             'payment': getparty.payment,
#             'current_date': getparty.current_date,
#             'additionalfield1': getparty.additionalfield1,
#             'additionalfield2': getparty.additionalfield2,
#             'additionalfield3': getparty.additionalfield3,
#         }

#         getparty.party_name = request.POST.get('partyname', '')
#         # ... (rest of the fields)

#         # Check for changes
#         if old_data != {
#             'party_name': getparty.party_name,
#             'trn_no': getparty.trn_no,
#             'contact': getparty.contact,
#             'trn_type': getparty.trn_type,
#             'state': getparty.state,
#             'address': getparty.address,
#             'email': getparty.email,
#             'openingbalance': getparty.openingbalance,
#             'payment': getparty.payment,
#             'current_date': getparty.current_date,
#             'additionalfield1': getparty.additionalfield1,
#             'additionalfield2': getparty.additionalfield2,
#             'additionalfield3': getparty.additionalfield3,
#         }:
#             # Changes detected, update history

#             # Save the party changes
#             getparty.save()

#             existing_transaction = Transactions_party.objects.filter(party=getparty).first()

#             if existing_transaction:
#                 existing_transaction.trans_type = 'Opening Balance'
#                 existing_transaction.trans_number = getparty.trn_no
#                 existing_transaction.trans_date = getparty.current_date
#                 existing_transaction.total = getparty.openingbalance
#                 existing_transaction.balance = getparty.openingbalance
#                 existing_transaction.save()

#             else:
#                 trans = Transactions_party(
#                     user=request.user,
#                     trans_type='Opening Balance',
#                     trans_number=getparty.trn_no,
#                     trans_date=getparty.current_date,
#                     total=getparty.openingbalance,
#                     balance=getparty.openingbalance,
#                     party=getparty
#                 )

#                 if request.user.is_company:
#                     trans.company = request.user.company
#                 else:
#                     trans.company = request.user.employee.company

#                 trans.save()

#                 # Now that trans is created, create history entry
#                 history_entry = PartyTransactionHistory(
#                     party=getparty,
#                     Transactions_party=trans,
#                     action="UPDATED",
#                     # Update with the current timestamp
#                 )
#                 history_entry.save()

#         return redirect('view_party', id=getparty.id)

#     return render(request, 'edit_party.html', {'getparty': getparty, 'party': party_qs, 'usr': request.user})


def history_party(request, id):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)

    fparty = Party.objects.get(id=id)
    ftrans = Transactions_party.objects.get(party=fparty)
    hst = PartyTransactionHistory.objects.filter(party=id)

    context = {'party': party, 'hst': hst, 'ftrans': ftrans, 'usr': request.user, 'fparty': fparty}
    return render(request, 'partyhistory.html', context)







def deleteparty(request,id):
    if request.user.is_company:
      party = Party.objects.filter(company = request.user.company)
    else:
      party = Party.objects.filter(company = request.user.employee.company)

    party=Party.objects.get(id=id)
    party.delete()
    return redirect('party_list')
    








# def shareTransactionpartyToEmail(request, id):
#   if request.user.is_company:
#     party = Party.objects.filter(company=request.user.company)
#   else:
#     party = Party.objects.filter(company=request.user.employee.company)

#     if request.method == "POST":
#         try:
#             fparty = Party.objects.get(id=id)
#             ftrans = Transactions_party.objects.filter(party=fparty)
#             context = {'party': party, 'usr': request.user, 'fparty': fparty, 'ftrans': ftrans}

#             email_message = request.POST.get('email_message')
#             my_subject = "Transaction REPORT"
#             emails_string = request.POST.get('email_ids')
#             emails_list = [email.strip() for email in emails_string.split(',')]

#             html_message = render_to_string('transaction_pdf.html', context)
#             plain_message = strip_tags(html_message)

#             pdf_content = BytesIO()
#             pisa.CreatePDF(html_message.encode("UTF-8"), dest=pdf_content)
#             pdf_content.seek(0)

#             filename = f'transaction.pdf'
#             message = EmailMultiAlternatives(
#                 subject=my_subject,
#                 body=f"Hi,\nPlease find the attached Transaction Report - \n{email_message}\n--\nRegards,\n",
#                 from_email='altostechnologies6@gmail.com',  # Update with your email
#                 to=emails_list,
#             )

#             message.attach(filename, pdf_content.read(), 'application/pdf')
#             message.send()

#             return HttpResponse('<script>alert("Report has been shared successfully!");window.location="/party_list"</script>')
#         except Exception as e:
#                 # Handle the exception, log the error, or provide an error message
#             return HttpResponse('<script>alert("Failed to send email!");window.location="/party_list"</script>')

#     return HttpResponse('<script>alert("Invalid Request!");window.location="/party_list"</script>')



from django.contrib import messages

def html_to_pdf(request):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)

    fparty = None
    ftrans = None

    if request.method == "POST":
        party_id = request.POST.get('party_id')
        try:
            fparty = Party.objects.get(id=party_id)
            ftrans = Transactions_party.objects.filter(party=fparty)
        except Party.DoesNotExist:
            messages.error(request, 'Party not found!')
            return HttpResponse('<script>window.location="/party_list"</script>')

    # Debug print/log
    print("Party:", party)
    print("User:", request.user)
    print("Fetched Party:", fparty)
    print("Fetched Transactions:", ftrans)

    context = {'party': party, 'usr': request.user, 'fparty': fparty, 'ftrans': ftrans}
    html_content = render(request, 'transaction_pdf.html', context).content

    # Create a BytesIO buffer to receive the PDF content
    pdf_buffer = BytesIO()

    # Use xhtml2pdf to generate the PDF
    pisa.CreatePDF(html_content, dest=pdf_buffer)

    # Move the buffer's cursor to the beginning
    pdf_buffer.seek(0)

    # Set the response content type
    response = HttpResponse(content_type='application/pdf')

    # Set the content-disposition header to force download
    response['Content-Disposition'] = 'filename="Transaction details.pdf"'

    # Write the PDF content to the response
    response.write(pdf_buffer.read())

    return response



from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse
from django.utils.html import strip_tags
from xhtml2pdf import pisa
from io import BytesIO

def shareTransactionpartyToEmail(request, id):
    if request.user.is_company:
        party = Party.objects.filter(company=request.user.company)
    else:
        party = Party.objects.filter(company=request.user.employee.company)
        
    if request.method == "POST":
        try:
            fparty = Party.objects.get(id=id)
            ftrans = Transactions_party.objects.filter(party=fparty)
            context = {'party': party, 'usr': request.user, 'fparty': fparty, 'ftrans': ftrans}
            
            email_message = request.POST.get('email_message')
            my_subject = "Transaction REPORT"
            emails_string = request.POST.get('email_ids')
            emails_list = [email.strip() for email in emails_string.split(',')]
            
            html_message = render_to_string('transaction_pdf.html', context)
            plain_message = strip_tags(html_message)
            
            pdf_content = BytesIO()
            pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content)
            pdf_content.seek(0)

            filename = f'transaction.pdf'
            message = EmailMultiAlternatives(
                subject=my_subject,
                body=f"Hi,\nPlease find the attached Transaction Report - \n{email_message}\n--\nRegards,\n",
                from_email='altostechnologies6@gmail.com',  # Update with your email
                to=emails_list,
            )

            message.attach(filename, pdf_content.read(), 'application/pdf')
            message.send()

            return HttpResponse('<script>alert("Report has been shared successfully!");window.location="/party_list"</script>')
        except Party.DoesNotExist:
            return HttpResponse('<script>alert("Party not found!");window.location="/party_list"</script>')
        except Exception as e:
            # Handle the exception, log the error, or provide an error message
            return HttpResponse(f'<script>alert("Failed to send email: {str(e)}");window.location="/party_list"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/party_list"</script>')





# def itemdetailinvoice(request):
#   if request.user.is_company:
#       company = request.user.company
#       parties = Party.objects.filter(company=company)
#   else:
#       company = request.user.employee.company
#       parties = Party.objects.filter(company=company)
#   itmid = request.GET['id']
#   itm = Item.objects.get(id=itmid)
#   hsn = itm.itm_hsn
#   price = itm.itm_sale_price
#   return JsonResponse({'hsn':hsn, 'price':price}) 

from django.http import JsonResponse

def itemdetailinvoice(request):
    try:
        if request.user.is_company:
            company = request.user.company
            parties = Party.objects.filter(company=company)
        else:
            company = request.user.employee.company
            parties = Party.objects.filter(company=company)

        itmid = request.GET.get('id')  # Using get to avoid KeyError
        itm = Item.objects.get(id=itmid)

        hsn = itm.itm_hsn
        price = itm.itm_sale_price

        return JsonResponse({'hsn': hsn, 'price': price,'parties':parties})
    except Item.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)







from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from .models import Party, Item, SalesInvoice

@login_required
def add_salesinvoice(request):
    try:
        if request.user.is_company:
            company = request.user.company
            parties = Party.objects.filter(company=company)
        else:
            company = request.user.employee.company
            parties = Party.objects.filter(company=company)
            
     
        
        items = Item.objects.filter(company=company)
        
        
        sales_invoices = SalesInvoice.objects.filter(company=company)
        if sales_invoices.exists():
            invoice_count = sales_invoices.last().invoice_no
            next_count = invoice_count + 1
        else:
            next_count = 1


        return render(request, 'add_salesinvoice.html', {'parties': parties, 'usr': request.user, 'count': next_count, 'items': items})

    except ObjectDoesNotExist:
        # Handle the case where the Company or Party object is not found
        return HttpResponse("Error: Company or Party not found.")






def party_details(request, party_name):
    try:
        party = Party.objects.get(party_name=party_name)
        data = {
            'contact': party.contact,
            'address': party.address,
            'openingbalance': party.openingbalance,
            'payment': party.payment,
        }
        return JsonResponse(data)
    except Party.DoesNotExist:
        return JsonResponse({'error': 'Party not found'},status=404)

from django.utils.dateparse import parse_date

# def salesinvoice_save_parties(request):
#     if request.method == 'POST':
#         if request.user.is_company:
#             company = request.user.company
#             parties = Party.objects.filter(company=company)
#         else:
#             company = request.user.employee.company
#             parties = Party.objects.filter(company=company)
        
#         party_name = request.POST['partyname']
#         trn_no = request.POST.get('trn_no', '')
#         contact = request.POST.get('contact', '')
#         trn_type = request.POST.get('trn_type', '')
#         state = request.POST.get('state', '')
#         address = request.POST.get('address', '')
#         email = request.POST.get('email', '')
#         opening_stock = request.POST['opening_stock']
#         at_price = request.POST['at_price']
#         openingbalance = request.POST.get('balance', '')
#         payment = request.POST.get('paymentType', '')
#         current_date_str = request.POST.get('currentdate', '')

#         # Validate current_date
#         if not current_date_str:
#             return render(request, 'add_salesinvoice.html', {'error_message': 'Current date is required'})
#         try:
#             current_date = parse_date(current_date_str)
#         except ValueError:
#             return render(request, 'add_salesinvoice.html', {'error_message': 'Invalid date format for current date'})

#         End_date = request.POST.get('enddate', None)
#         additionalfield1 = request.POST.get('additionalfield1', '')
#         additionalfield2 = request.POST.get('additionalfield2', '')
#         additionalfield3 = request.POST.get('additionalfield3', '')

#         # Check if required fields are missing
#         if not party_name:
#             return render(request, 'add_salesinvoice.html', {'error_message': 'Party name is required'})

#         part = Party(
#             party_name=party_name, 
#             trn_no=trn_no, 
#             contact=contact, 
#             trn_type=trn_type, 
#             state=state,
#             address=address, 
#             email=email, 
#             openingbalance=openingbalance,
#             opening_stock=opening_stock, 
#             at_price=at_price, 
#             payment=payment,
#             current_date=current_date, 
#             End_date=End_date, 
#             additionalfield1=additionalfield1,
#             additionalfield2=additionalfield2, 
#             additionalfield3=additionalfield3, 
#             company=company
#         )
#         part.save() 

#         return render(request,'add_salesinvoice.html')

#     return render(request, 'add_salesinvoice.html')
from django.core.exceptions import ValidationError
from django.utils.datastructures import MultiValueDictKeyError

def salesinvoice_save_parties(request):
    if request.user.is_company:
        cmp = request.user.company
    else:
        cmp = request.user.employee.company

    error_message = None  # Initialize error message

    if request.method == 'POST':
        try:
            party_name = request.POST['partyname']
            trn_no = request.POST['trn_no']
            contact = request.POST.get('contact', '')

            trn_type = request.POST.get('trn_type', '')

            state = request.POST.get('state', '')
            address = request.POST.get('address', '')

            email = request.POST.get('email', '')
            opening_stock = request.POST['opening_stock']
            at_price = request.POST['at_price']
            openingbalance = request.POST.get('balance', '')
            payment = request.POST.get('paymentType', '')
            current_date = request.POST['currentdate']

            End_date = request.POST.get('enddate', None)
            additionalfield1 = request.POST.get('additionalfield1', '')

            additionalfield2 = request.POST.get('additionalfield2', '')

            additionalfield3 = request.POST.get('additionalfield3', '')
        except MultiValueDictKeyError:
            # Handle missing keys in request.POST
            error_message = 'Required fields are missing.'
            return render(request, 'add_salesinvoice.html', {'error_message': error_message})

        try:
            # Check if a party with the same trn_no already exists
            if Party.objects.filter(trn_no=trn_no, company=cmp).exists():
                error_message = 'An error occurred while processing your request. TRN number already exists. Please enter a unique TRN number.'
            # Check if a party with the same email already exists
            elif Party.objects.filter(email=email, company=cmp).exists():
                error_message = 'An error occurred while processing your request. Email already exists. Please enter a unique email address.'
            else:
                part = Party(party_name=party_name, trn_no=trn_no, contact=contact, trn_type=trn_type, state=state,
                             address=address, email=email, openingbalance=openingbalance, opening_stock=opening_stock,
                             at_price=at_price, payment=payment, current_date=current_date, End_date=End_date,
                             additionalfield1=additionalfield1, additionalfield2=additionalfield2,
                             additionalfield3=additionalfield3, company=cmp)
                part.save()

                trans = Transactions_party(user=request.user, trans_type='Opening Balance', trans_number=trn_no,
                                           trans_date=current_date, total=openingbalance, balance=openingbalance,
                                           party=part, company=cmp)
                trans.save()

                tr_history = PartyTransactionHistory(party=part, Transactions_party=trans, action="CREATED")
                tr_history.save()

                return redirect('add_salesinvoice')

        except ValidationError as e:
            error_message = str(e)

    return render(request, 'add_salesinvoice.html', {'error_message': error_message})



    

# def itemdata_salesinvoiceedit(request):
#   itmid = request.GET['id']
#   print(itmid)
#   itm = Item.objects.get(id=itmid)
#   print(itm)
#   hsn = itm.itm_hsn
#   # gst = itm.itm_gst
#   # igst = itm.itm_igst
#   price = itm.itm_sale_price
#   qty = itm.itm_at_price
#   return JsonResponse({'hsn':hsn, 'price':price, 'qty':qty})
    








from django.http import Http404

def itemdata_salesinvoiceedit(request):
  if request.user.is_company:
            company = request.user.company
            parties = Party.objects.filter(company=company)
  else:
            company = request.user.employee.company
            parties = Party.objects.filter(company=company)
  try:
        
        itmid = request.GET['id']
        itm = Item.objects.get(id=itmid)
        hsn = itm.itm_hsn
        vat = itm.itm_vat
        # gst = itm.itm_gst
        # igst = itm.itm_igst
        price = itm.itm_sale_price
        qty = itm.itm_at_price
        return JsonResponse({'hsn': hsn, 'price': price, 'qty': qty, 'vat':vat})
  except Item.DoesNotExist:
        raise Http404("Item not found")


def save_sales_invoice(request):
  if request.user.is_company:
      company = request.user.company
      parties = Party.objects.filter(company=company)
  else:
    company = request.user.employee.company
    parties = Party.objects.filter(company=company)
  
  
  
  
  if request.method == 'POST':      
    party_name = request.POST.get('partyname')
    contact = request.POST.get('contact')
    address = request.POST.get('address')
    invoice_no = request.POST.get('invoiceno')
    tax=request.POST.get('tax')
    date = request.POST.get('date')
    product = tuple(request.POST.getlist("product[]"))
    hsn =  tuple(request.POST.getlist("hsn[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    rate =  tuple(request.POST.getlist("price[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    vat =  tuple(request.POST.getlist("vat[]"))
    total =  tuple(request.POST.getlist("total[]"))
    description = request.POST.get('description')
    subtotal = float(request.POST.get('subtotal'))
    adjust = request.POST.get("adj")
    grandtotal=request.POST.get('grandtotal')
    taxamount = request.POST.get("taxamount")
    party=Party.objects.get(party_name=party_name)

        
      
    sales_invoice = SalesInvoice(
        company=company,
        party_name=party_name,
        contact=contact,
        address=address,
        invoice_no=invoice_no,
        date=date,
        description=description,
        subtotal=subtotal,
        vat=request.POST['vat'],
        total_taxamount=taxamount if taxamount is not None else 0,  # Provide a default value if taxamount is None
        adjustment=adjust,
        grandtotal=grandtotal,
        
    )

    sales_invoice.save()

    tr_history = SalesInvoiceTransactionHistory(company=company,
                                                 
                                          salesinvoice=sales_invoice,
                                          action="CREATED",
                                          done_by_name=party.party_name,
                                          )
    tr_history.save()

    invoice = SalesInvoice.objects.get(id=sales_invoice.id)
    mapped = []  # Initialize mapped
    print(product,hsn,qty,discount,vat,total,rate)
    if len(product)==len(hsn)==len(qty)==len(rate)==len(discount)==len(vat)==len(total):
      mapped=zip(product, hsn, qty, rate, discount, vat, total)
      mapped=list(mapped)
    for ele in mapped: 
      itm = Item.objects.get(id=ele[0])
      SalesInvoiceItem.objects.create(item=itm, hsn=ele[1], quantity=ele[2], rate=ele[3], discount=ele[4], tax=ele[5], totalamount=ele[6], salesinvoice=invoice, company=company)

    if 'save_and_new' in request.POST:
      return redirect(add_salesinvoice)
    else:
      return redirect('salesinvoice_billtemplate', sales_invoice.id)

  
  return HttpResponse("Default response")
        # if 'save_and_new' in request.POST:
        #   return redirect(add_salesinvoice)
        # else:
        #   return redirect('salesinvoice_billtemplate',sales_invoice.id)
    # return render(request, 'add_salesinvoice.html')

  

    #     if taxamount is not None and taxamount != "":
    #         sales_invoice = SalesInvoice(
    #             company=company,
    #             # party=party_instance,
    #             parties=parties,
    #             party_name=party_name,
    #             contact=contact,
    #             address=address,
    #             invoice_no=invoice_no,
    #             date=date,
    #             description=description,
    #             subtotal=subtotal,
    #             vat=vat,
    #             total_taxamount=taxamount,
    #             adjustment=adjust,
    #             grandtotal=grandtotal,
    #         )
    #         sales_invoice.save()


    
    #         if sales_invoice is not None:
    #             tr_history = SalesInvoiceTransactionHistory(
    #                 company=company,
    #                 parties=parties,
    #                 salesinvoice=sales_invoice,
    #                 action="CREATED",
    #                 done_by_name=parties.name,
    #             )
    #             tr_history.save()

    #         try:
    #             # Attempt to fetch the SalesInvoice object again
    #             invoice = SalesInvoice.objects.get(id=sales_invoice.id)
    #             mapped = []  # Initialize mapped
    #             if len(product) == len(hsn) == len(qty) == len(rate) == len(discount) == len(tax) == len(total):
    #                 mapped = zip(product, hsn, qty, rate, discount, tax, total)
    #                 mapped = list(mapped)
    #                 for ele in mapped:
    #                     itm = Item.objects.get(id=ele[0])
    #                     SalesInvoiceItem.objects.create(
    #                         item=itm,
    #                         hsn=ele[1],
    #                         quantity=ele[2],
    #                         rate=ele[3],
    #                         discount=ele[4],
    #                         tax=ele[5],
    #                         totalamount=ele[6],
    #                         salesinvoice=invoice,
    #                         company=company,
    #                     )
    #         except SalesInvoice.DoesNotExist:
    #             # Handle the case where SalesInvoice.DoesNotExist exception is raised
    #             return HttpResponse("Error: Sales Invoice not found", status=500)
    #         except Exception as e:
    #             # Handle other exceptions that might occur during the fetch
    #             return HttpResponse(f"Error: {e}", status=500)
    #     else:
    #         # Handle the case where sales_invoice was not saved successfully
    #         return HttpResponse("Error: Sales Invoice not saved successfully", status=500)



        

    # if 'save_and_new' in request.POST:
    #     return redirect(add_salesinvoice)
    # else:
    #     if sales_invoice:
    #         return redirect('salesinvoice_billtemplate', sales_invoice.id)
    #     else:
    #         return HttpResponse("Error: Sales Invoice not saved successfully", status=500)

from django.urls import reverse


def item_save_invoice(request):
    if request.method == 'POST':
        itm_type = request.POST.get('itm_type')
        if itm_type:
            item_type = 'Service'
        else:
            item_type = 'Goods'

        itm_name = request.POST.get('name')
        itm_hsn = request.POST.get('hsn')
        itm_unit = request.POST.get('unit')
        itm_vat = request.POST.get('vat')
        taxable_result = request.POST.get('taxable_result')
        itm_sale_price = request.POST.get('sale_price')
        itm_purchase_price = request.POST.get('purchase_price')
        stock_in_hand = request.POST.get('stock_in_hand') or 0
        itm_at_price = request.POST.get('at_price') or 0
        itm_date = request.POST.get('itm_date')
        


        item = Item(
            user=request.user,
            itm_type=item_type,
            itm_name=itm_name,
            itm_hsn=itm_hsn,
            itm_unit=itm_unit,
            itm_vat=itm_vat,
            itm_taxable=taxable_result,
            itm_sale_price=itm_sale_price,
            itm_purchase_price=itm_purchase_price,
            itm_stock_in_hand=stock_in_hand,
            itm_at_price=itm_at_price,
            itm_date=itm_date
        )

        if request.user.is_company:
            item.company = request.user.company
        else:
            item.company = request.user.employee.company

        item.save()

        return JsonResponse({'success': True})





def item_invoicedropdown(request):
  if request.user.is_company:
    company = request.user.company
    parties = Party.objects.filter(company=company)
  else:
    company = request.user.employee.company
    parties = Party.objects.filter(company=company)
  print(parties)
  options = {}
  option_objects = Item.objects.filter(company=company)
  for option in option_objects:
      options[option.id] = [option.itm_name]
  return JsonResponse(options)
  



def salesinvoice_billtemplate(request, id):
    if request.user.is_company:
        company = request.user.company
        parties = Party.objects.filter(company=company)
    else:
        company = request.user.employee.company
        parties = Party.objects.filter(company=company)

    print(parties)
  
    history = SalesInvoiceTransactionHistory.objects.filter(salesinvoice=id)
    salesinvoice = SalesInvoice.objects.get(id=id)

    # Assuming SalesInvoiceItem has a ForeignKey relationship with SalesInvoice
    salesinvoiceitems = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)

    dis = 0
    for itm in salesinvoiceitems:
        dis += int(itm.discount)

    itm_len = len(salesinvoiceitems)

    return render(request, 'salesinvoice_billtemplate.html', {
        'parties': parties,
        'company': company,
        'history': history,
        'salesinvoice': salesinvoice,
        'salesinvoiceitems': salesinvoiceitems,
        'dis': dis,
        'itm_len': itm_len,
    })



def view_salesinvoice(request):
    if request.user.is_company:
        company = request.user.company
        parties = Party.objects.filter(company=company)
    else:
        company = request.user.employee.company
        parties = Party.objects.filter(company=company)

    print(parties)
    
    
    party = Party.objects.filter(company=company)
    item = Item.objects.filter(company=company)
    
    salesinvoice = SalesInvoice.objects.filter(company=company)
    for i in salesinvoice:
        last_transaction = SalesInvoiceTransactionHistory.objects.filter(salesinvoice=i).last()
        if last_transaction:
            i.action = last_transaction.action
            i.done_by_name = last_transaction.done_by_name
        else:
            i.action = None
            i.done_by_name = None

    return render(request, 'view_salesinvoice.html', {'parties':parties,'party': party, 'item': item, 'salesinvoice': salesinvoice,'company':company})


def deletesalesinvoice(request,id):
    salesinvoice=SalesInvoice.objects.get(id=id)
    salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)
    salesinvoice.delete()
    salesinvoiceitem.delete()
    return redirect('view_salesinvoice')






def edit_salesinvoice(request,id):
  if request.user.is_company:
    company = request.user.company
    parties = Party.objects.filter(company=company)
  else:
    company = request.user.employee.company
    parties = Party.objects.filter(company=company)

  print(parties)
  
  getinoice=SalesInvoice.objects.get(id=id)
  getitem=SalesInvoiceItem.objects.filter(salesinvoice=id)
  party=Party.objects.filter(company=company)
  item=Item.objects.filter(company=company)
  


  return render(request, 'edit_salesinvoice.html',{'parties':parties,'getinoice':getinoice,'getitem':getitem,'party':party,'item':item})


def editsave_salesinvoice(request,id):

  if request.user.is_company:
    company = request.user.company
    parties = Party.objects.filter(company=company)
  else:
    company = request.user.employee.company
    parties = Party.objects.filter(company=company)


    sales_invoice=SalesInvoice.objects.get(id=id,company=company,parties=parties)
    
    sales_invoice.party_name = request.POST.get('partyname')
    sales_invoice.contact = request.POST.get('contact')
    sales_invoice.address = request.POST.get('address')
    sales_invoice.invoice_no = request.POST.get('invoiceno')
    sales_invoice.date = request.POST.get('date')
    sales_invoice.description = request.POST.get('description')
    sales_invoice.subtotal =float(request.POST.get('subtotal'))
    sales_invoice.vat = request.POST.get('vat')
    sales_invoice.total_taxamount = request.POST.get('taxamount')
    sales_invoice.adjustment = request.POST.get('adj')
    sales_invoice.grandtotal = request.POST.get('grandtotal')
    
    sales_invoice.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    vat =tuple( request.POST.getlist("vat[]"))
    discount = tuple(request.POST.getlist("discount[]"))
    total = tuple(request.POST.getlist("total[]"))
    SalesInvoiceItem.objects.filter(salesinvoice=sales_invoice,company=company).delete()
    if len(product)==len(qty)==len(qty)==len(vat):
      mapped=zip(product,qty,vat,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = Item.objects.get(id=ele[0])
        SalesInvoiceItem.objects.create(item =itm,quantity=ele[1], tax=ele[2],discount=ele[3],totalamount=ele[4],salesinvoice=sales_invoice,company=company)

    tr_history = SalesInvoiceTransactionHistory(company=company,
                                              
                                          salesinvoice=sales_invoice,
                                          action="UPDATED",
                                          done_by_name=parties.name,
                                          )
    tr_history.save()

    
    return redirect('view_salesinvoice')

  return render(request, 'edit_salesinvoice.html')





from django.shortcuts import render, get_object_or_404
from .models import Party, Employee, SalesInvoiceTransactionHistory

def salesinvoicehistory(request, id):
    if request.user.is_company:
        company = request.user.company
        employee = None  # Initialize employee variable
    else:
        # Use get_object_or_404 to handle DoesNotExist exception
        employee = get_object_or_404(Employee, user=request.user)
        company = employee.company

    parties = Party.objects.filter(company=company)
    emp = employee  # Use the retrieved employee object directly

    history = SalesInvoiceTransactionHistory.objects.filter(salesinvoice=id)

    return render(request, 'salesinvoicehistory.html', {
        'parties': parties,
        'history': history,
        'company': company,
        'emp': emp,
    })




# from django.http import JsonResponse

# from django.db.models import Sum

# def profit_loss_data(request):
    
#   if request.user.is_company:
#     company = request.user.company
#     parties = Party.objects.filter(company=company)
#   else:
#     company = request.user.employee.company
#     parties = Party.objects.filter(company=company)
#     # staff = staff_details.objects.get(id=staff_id)
#     # company_instance = company.objects.get(id=staff.company.id)
#   labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sept", "Oct", "Nov", "Dec"]


#   sales_data = (
#       SalesInvoice.objects.filter(date__year=2024,company=company)
#       .values('date__month')
#       .annotate(grandtotal_sum=Sum('grandtotal'))
#   )

#   # Create a dictionary with monthly sales data
#   sales_dict = {item['date__month']: item['grandtotal_sum'] for item in sales_data}

#   # Fill in sales values for each month
#   sales = [sales_dict.get(month, 0) for month in range(1, 13)]

#   data = {'labels': labels, 'sales': sales, 'parties':parties}
#   return JsonResponse(data)
  



# def graph_salesinvoice(request):
#     if request.user.is_company:
#         company = request.user.company
#         parties = Party.objects.filter(company=company)
#     else:
#         company = request.user.employee.company
#         parties = Party.objects.filter(company=company)

#     # Assuming parties is a queryset, you might want to get a specific party from it
#     # For example, you could get the first party in the queryset
#     party = parties.first()

#     # Check if a party exists before proceeding
#     if party:
#         cmp = Company.objects.get(id=party.company.id)
#         user = cmp.user  # Access user attribute of the Company model
#         salesinvoice = SalesInvoiceItem.objects.filter(company=cmp)

#         return render(request, 'graph_salesinvoice.html', {'salesinvoice': salesinvoice, 'user': user})

#     # Handle the case where no party is found
#     # For example, return an error response or redirect to another page
#     return HttpResponse("No party found.")

from django.core.serializers import serialize
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum

from django.http import JsonResponse
from django.http import JsonResponse

def profit_loss_data(request):
    try:
        company = get_user_company(request.user)
        print(f"Company: {company}")

        parties = Party.objects.filter(company=company)
        print(f"Parties: {parties}")

        labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sept", "Oct", "Nov", "Dec"]

        sales_data = (
            SalesInvoice.objects.filter(date__year=2024, company=company)
            .values('date__month')
            .annotate(grandtotal_sum=Sum('grandtotal'))
        )

        print(f"Sales Data: {sales_data}")

        sales_dict = {item['date__month']: item['grandtotal_sum'] for item in sales_data}
        sales = [sales_dict.get(month, 0) for month in range(1, 13)]

        data = {'labels': labels, 'sales': sales, 'parties': parties}
        return JsonResponse(data)
    except Exception as e:
        # Log the exception details
        print(f"Exception in profit_loss_data: {e}")
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


def get_user_company(user):
    if user.is_company:
        return user.company
    elif user.employee:
        return user.employee.company
    # Handle the case where neither company nor employee is found
    raise ValueError("User is not associated with a company or employee.")

def graph_salesinvoice(request):
    company = get_user_company(request.user)
    parties = Party.objects.filter(company=company)

    party = parties.first()

    if party:
        user = company.user  # Assuming user is an attribute of the Company model
        salesinvoice = SalesInvoiceItem.objects.filter(company=company)
        return render(request, 'graph_salesinvoice.html', {'salesinvoice': salesinvoice, 'user': user})

    return HttpResponse("No party found.")



from openpyxl import Workbook
from django.http import HttpResponse

from openpyxl import load_workbook
from django.contrib import messages
from django.utils import timezone



def importsalesinvoice_excel(request):
    if request.method == 'POST' and request.FILES['billfile'] and request.FILES['prdfile']:
        if request.user.is_company:
          company = request.user.company
          parties = Party.objects.filter(company=company)
        else:
          company = request.user.employee.company
          parties = Party.objects.filter(company=company)

        
        totval = int(SalesInvoice.objects.filter(company=company).last().invoice_no) + 1

        excel_bill = request.FILES['billfile']
        excel_b = load_workbook(excel_bill)
        eb = excel_b['Sheet1']
        excel_prd = request.FILES['prdfile']
        excel_p = load_workbook(excel_prd)
        ep = excel_p['Sheet1']

        for row_number1 in range(2, eb.max_row + 1):
            billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
            part = Party.objects.get(party_name=billsheet[0], email=billsheet[1], company=company)
            SalesInvoice.objects.create(party=part,
                                        date=billsheet[2],
                                        state_of_supply=billsheet[3],
                                        invoice_no=totval,
                                        company=company, parties=parties)
            invoice = SalesInvoice.objects.create(
                        party=part,
                        date=billsheet[2],
                        state_of_supply=billsheet[3],
                        invoice_no=totval,
                        company=company
                    )


            invoice.paymenttype = billsheet[4]
                    
            invoice.save()

            subtotal = total_taxamount = 0
            for row_number2 in range(2, ep.max_row + 1):
                        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
                        if prdsheet[0] == row_number1:
                            itm = ItemModel.objects.get(item_name=prdsheet[1], item_hsn=prdsheet[2], company=company_instance)
                            total = int(prdsheet[3]) * int(itm.item_sale_price) - int(prdsheet[4])

                            # Create SalesInvoiceItem object
                            SalesInvoiceItem.objects.create(
                                salesinvoice=invoice,
                                company=company_instance,
                                item=itm,
                                staff=staff,
                                quantity=prdsheet[3],
                                discount=prdsheet[4]
                            )

                            tax = int(prdsheet[5])
                            subtotal += total
                            tamount = total * (tax / 100)
                            total_taxamount += tamount

        gtotal = subtotal + total_taxamount + float(billsheet[6])
        balance = gtotal - float(billsheet[7])
        gtotal = round(gtotal, 2)
        balance = round(balance, 2)

        invoice.subtotal = round(subtotal, 2)
        invoice.total_taxamount = round(total_taxamount, 2)
        invoice.adjustment = round(billsheet[6], 2)
        invoice.grandtotal = gtotal
        invoice.paidoff = round(billsheet[7], 2)
        invoice.totalbalance = balance
        invoice.save()

        SalesInvoiceTransactionHistory.objects.create(
            salesinvoice=invoice,
            staff=invoice.staff,
            company=invoice.company,
            action='Created',
            done_by_name=invoice.staff.first_name
        )

    return JsonResponse({'message': 'File uploaded successfully!'})
  # except Exception as e:
  #   return JsonResponse({'message': f'File upload failed: {str(e)}'})

  #   return JsonResponse({'message': 'File upload failed: Invalid request method or missing files.'})